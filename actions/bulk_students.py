import csv
import io
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import insert
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Student, utcnow


def _normalize_row(obj: dict[str, Any]) -> tuple[str, str] | None:
    """Returns (admission_id, name) or None if invalid."""
    keys = {k.lower().strip(): v for k, v in obj.items() if k is not None}
    aid = keys.get("id") or keys.get("admission") or keys.get("admission_no") or keys.get("student_id")
    name = keys.get("name") or keys.get("full_name")
    if aid is None or name is None:
        return None
    aid_s = str(aid).strip()
    name_s = str(name).strip()
    if not aid_s or not name_s:
        return None
    return aid_s, name_s


def parse_student_rows_json(payload: list[dict[str, Any]]) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for item in payload:
        t = _normalize_row(item)
        if t:
            out.append(t)
    return out


def parse_student_rows_csv(data: bytes) -> list[tuple[str, str]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    out: list[tuple[str, str]] = []
    for row in reader:
        t = _normalize_row(row)
        if t:
            out.append(t)
    return out


def parse_student_rows_xlsx(data: bytes) -> list[tuple[str, str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    header_l = [str(h).lower().strip() if h is not None else "" for h in header]
    try:
        idx_id = next(i for i, h in enumerate(header_l) if h in ("id", "admission", "admission_no", "student_id"))
        idx_name = next(i for i, h in enumerate(header_l) if h in ("name", "full_name"))
    except StopIteration:
        wb.close()
        return []

    out: list[tuple[str, str]] = []
    for row in rows_iter:
        if row is None:
            continue
        def cell(i: int):
            if i >= len(row):
                return None
            return row[i]

        aid = cell(idx_id)
        name = cell(idx_name)
        if aid is None and name is None:
            continue
        aid_s = str(aid).strip() if aid is not None else ""
        name_s = str(name).strip() if name is not None else ""
        if aid_s and name_s:
            out.append((aid_s, name_s))
    wb.close()
    return out


async def bulk_create_students(
    db: AsyncSession,
    *,
    school_id: str,
    class_id: str,
    rows: list[tuple[str, str]],
) -> int:
    if not rows:
        return 0
    seen: set[str] = set()
    mappings: list[dict[str, Any]] = []
    for admission_id, name in rows:
        if admission_id in seen:
            continue
        seen.add(admission_id)
        mappings.append(
            {
                "school_id": school_id,
                "id": admission_id,
                "name": name,
                "c_id": class_id,
                "is_deleted": False,
                "date_added": utcnow(),
            }
        )
    if not mappings:
        return 0

    await db.execute(insert(Student), mappings)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise
    return len(mappings)
